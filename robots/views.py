"""
Views for robots
"""
import json
from datetime import datetime, timedelta
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.dateparse import parse_datetime
from django.shortcuts import render
from django.db.models import Count
from openpyxl import Workbook
from .models import Robot

def create_robot_web(request):
    """
    Render form for creating a robot
    """
    return render(request, 'create_robot.html')

@csrf_exempt
def robot_create_view(request):
    """
    Handle POST requests to create a robot.

    Args:
    request (HttpRequest): The HTTP request containing JSON data.
    Returns:
    JsonResponse: A JSON response indicating the result of the operation.
    """
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            
            # Валидация обязательных полей
            required_fields = ['model', 'version', 'created']
            for field in required_fields:
                if field not in data:
                    return JsonResponse({'error': f'Missing field: {field}'}, status=400)
            
            # Дополнительная валидация
            model = data['model']
            version = data['version']
            is_available = data['is_available']
            created = parse_datetime(data['created'])
            
            if not created:
                return JsonResponse({'error': 'Invalid date format for "created". Use YYYY-MM-DD HH:MM:SS.'}, status=400)
            
            if not model.isalnum() or len(model) > 2:
                return JsonResponse({'error': 'Invalid "model". It must be alphanumeric and at most 2 characters.'}, status=400)
            
            if len(version) > 2:
                return JsonResponse({'error': 'Invalid "version". It must be at most 2 characters.'}, status=400)
            
            robot = Robot.objects.create( # pylint: disable=no-member
                serial=str(Robot.objects.count() + 1).zfill(5),  # pylint: disable=no-member
                model=model,
                version=version,
                is_available=is_available,
                created=created
            )
            return JsonResponse({'message': 'Robot created successfully', 'robot_id': robot.id}, status=201)
        
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
        except Exception as e: # pylint: disable=broad-except
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Only POST method is allowed'}, status=405)

def robot_list_view(request):
    """
    Handle GET requests to retrieve a list of robots.

    Args:
    request (HttpRequest): The HTTP request object.

    Returns:
    JsonResponse: A JSON response containing a list of robots.
    """
    if request.method == "GET":
        try:
            robots = Robot.objects.all() # pylint: disable=no-member

            robot_list = [
                {
                    'id': robot.id,
                    'serial': robot.serial,
                    'model': robot.model,
                    'version': robot.version,
                    'is_available': robot.is_available,
                    'created': robot.created.strftime('%Y-%m-%d %H:%M:%S')
                }
                for robot in robots
            ]

            return JsonResponse({'robots': robot_list}, status=200)

        except Exception as e: # pylint: disable=broad-except
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Only GET method is allowed'}, status=405)

def export_robots_summary(request):
    """
    Handle GET requests to export a summary of robots.

    Args:
    request (HttpRequest): The HTTP request object.

    Returns:
    HttpResponse: An Excel file containing a summary of robots.
    """
    if request.method == "GET":
        try:
            end_date = datetime.now()
            start_date = end_date - timedelta(days=7)

            robots = Robot.objects.filter(created__range=(start_date, end_date)) # pylint: disable=no-member

            # Сгруппировать данные по модели и версии, подсчитать количество
            summary = robots.values('model', 'version').annotate(count=Count('id'))

            # Создаем новый Excel-файл
            wb = Workbook()
            models = set(item['model'] for item in summary)
            
            for model in models:
                # Создаем лист для каждой модели
                ws = wb.create_sheet(title=model)
                ws.append(['Модель', 'Версия', 'Количество за неделю'])

                # Добавляем строки с данными
                for item in summary:
                    if item['model'] == model:
                        ws.append([item['model'], item['version'], item['count']])
            
            # Удаляем пустой лист, созданный по умолчанию
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="robots_summary.xlsx"'
            wb.save(response)

            return response

        except Exception as e: # pylint: disable=broad-except
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Only GET method is allowed'}, status=405)

@csrf_exempt
def robot_update_is_available(request, robot_id):
    """
    Handle PATCH requests to update the is_available field of a robot.

    Args:
    request (HttpRequest): The HTTP request object.
    robot_id (int): The ID of the robot to update.

    Returns:
    JsonResponse: A JSON response indicating the success or failure of the update.
    """
    if request.method == "PATCH":
        try:
            robot = Robot.objects.get(id=robot_id) # pylint: disable=no-member

            data = json.loads(request.body)
            is_available = data.get("is_available")

            if not isinstance(is_available, bool):
                return JsonResponse({'error': '"is_available" must be a boolean value (true/false).'}, status=400)

            robot.is_available = is_available
            robot.save()

            return JsonResponse({
                'message': 'is_available updated successfully',
                'robot_id': robot.id,
                'is_available': robot.is_available
            }, status=200)

        except Robot.DoesNotExist: # pylint: disable=no-member
            return JsonResponse({'error': 'Robot not found'}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
        except Exception as e: # pylint: disable=broad-except
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Only PATCH method is allowed'}, status=405)
